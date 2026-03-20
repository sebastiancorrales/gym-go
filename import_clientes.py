"""
Script de importación de clientes - OLYMPO GYM
Lee REGISTROS OLYMPO GYM.xlsx e inserta directamente en la DB de gym-go.

Uso:
    python import_clientes.py [--dry-run]
"""

import sys, uuid, sqlite3, re
import openpyxl
from datetime import datetime, timedelta

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

DB_PATH   = r"c:\Users\sebas\OneDrive\Desktop\gym-go\gym-go.db"
XLSX_PATH = r"C:\Users\sebas\Downloads\REGISTROS OLYMPO GYM.xlsx"
DRY_RUN   = "--dry-run" in sys.argv
GYM_ID    = "00000000-0000-0000-0000-000000000000"

PLAN_ALIAS = {
    "MENSUAL": "MENSUAL", "MENSUALIDAD": "MENSUAL",
    "QUINCENA": "QUINCENA", "QUINCE3NA": "QUINCENA",
    "SEMANA": "SEMANA",
    "PAREJA": "PAREJA", "PAREJAS": "PAREJA",
    "PANA": "PANAS", "PANAS": "PANAS",
    "TRIO": "TRIO", "TRÍO": "TRIO",
    "CUARTETO": "CUARTETO",
}

PLANES_BASE = {
    "MENSUAL":   (30,  70000, 1),
    "QUINCENA":  (15,  40000, 1),
    "SEMANA":    (7,   30000, 1),
    "PAREJA":    (30, 130000, 2),
    "PANAS":     (30, 130000, 3),
    "TRIO":      (30, 180000, 3),
    "CUARTETO":  (28, 240000, 4),
}

# ── Sanitización ──────────────────────────────────────────────────────────────

def clean_name(s):
    """Limpia un nombre: quita saltos, espacios extra, capitaliza."""
    if not s: return ""
    s = str(s).replace('\n', ' ').replace('\r', ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    # Quitar caracteres raros que no sean letras/espacios/acentos
    s = re.sub(r'[^\w\s\-\.]', '', s, flags=re.UNICODE)
    return s.title()

def clean_doc(s):
    """Extrae solo dígitos de una cédula (soporta cédulas con texto mezclado)."""
    if s is None: return ""
    s = str(s).strip()
    # Quitar .0 de números flotantes
    s = re.sub(r'\.0$', '', s)
    # Extraer solo dígitos si hay texto mezclado
    digits = re.sub(r'[^\d]', '', s)
    # Si la cadena original era solo dígitos, usarla; si hay basura, usar los dígitos
    if re.match(r'^\d+$', s):
        return s[:15]
    return digits[:15] if len(digits) >= 4 else ""

def clean_phone(s):
    if s is None: return ""
    s = str(s).strip().replace('.0', '').replace(' ', '')
    return re.sub(r'[^\d\+]', '', s)[:20]

def split_multiline(val):
    """Divide un campo de celda multilinea en lista limpia."""
    if val is None: return []
    return [x.strip() for x in str(val).split('\n') if x.strip()]

def new_uuid(): return str(uuid.uuid4())
def now_iso(): return datetime.now().isoformat()

def to_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try: return datetime.strptime(val.strip(), fmt)
            except: pass
    return None

def add_duration(start, days):
    days = int(days)
    if days == 365: return start.replace(year=start.year + 1)
    if days % 30 == 0:
        import calendar
        months = days // 30
        m = start.month - 1 + months
        year = start.year + m // 12
        month = m % 12 + 1
        day = min(start.day, calendar.monthrange(year, month)[1])
        return start.replace(year=year, month=month, day=day)
    return start + timedelta(days=days)

# ── DB ────────────────────────────────────────────────────────────────────────

con = sqlite3.connect(DB_PATH)
con.row_factory = sqlite3.Row
cur = con.cursor()

def get_or_create_plan(plan_norm):
    for alias in [plan_norm.lower(), plan_norm.upper(), plan_norm.capitalize()]:
        cur.execute("SELECT id FROM plans WHERE LOWER(name) = ? AND gym_id = ?",
                    (alias.lower(), GYM_ID))
        row = cur.fetchone()
        if row: return row["id"]

    duration_days, price, max_members = PLANES_BASE.get(plan_norm, (30, 70000, 1))
    plan_id = new_uuid()
    ahora = now_iso()
    if not DRY_RUN:
        cur.execute("""
            INSERT INTO plans (id, gym_id, name, duration_days, price, enrollment_fee,
                               max_members, color, display_order, is_featured, status, created_at, updated_at)
            VALUES (?,?,?,?,?,0,?,'#3B82F6',0,0,'ACTIVE',?,?)
        """, (plan_id, GYM_ID, plan_norm.capitalize(), duration_days, price,
              max_members, ahora, ahora))
    print(f"  + Plan creado: {plan_norm} ({duration_days}d, ${price:,}, max={max_members})")
    return plan_id

def get_or_create_user(first_name, last_name, doc, phone):
    first_name = clean_name(first_name)
    last_name  = clean_name(last_name)
    doc        = clean_doc(doc)
    phone      = clean_phone(phone)

    if not first_name: first_name = "Sin"
    if not last_name:  last_name  = "Nombre"

    if doc:
        cur.execute("SELECT id FROM users WHERE document_number = ? AND gym_id = ?", (doc, GYM_ID))
        row = cur.fetchone()
        if row: return row["id"], False

    user_id = new_uuid()
    email = f"member.{doc or user_id[:8]}@gymgo.internal"
    ahora = now_iso()
    if not DRY_RUN:
        cur.execute("""
            INSERT INTO users (id, gym_id, email, password_hash, first_name, last_name,
                               document_number, phone, role, status, email_verified,
                               qr_code, created_at, updated_at)
            VALUES (?,?,?,'auto-import',?,?,?,?,'MEMBER','ACTIVE',0,?,?,?)
        """, (user_id, GYM_ID, email, first_name, last_name,
              doc or None, phone or None, user_id[:8], ahora, ahora))
    return user_id, True

def create_subscription(user_id, plan_id, start_date, days, price, is_primary=True):
    if not start_date: start_date = datetime.now()
    end_date = add_duration(start_date, days)
    status   = "EXPIRED" if end_date < datetime.now() else "ACTIVE"
    total    = price if is_primary else 0
    sub_id   = new_uuid()
    ahora    = now_iso()
    if not DRY_RUN:
        cur.execute("""
            INSERT INTO subscriptions (id, user_id, plan_id, gym_id, start_date, end_date,
                price_paid, enrollment_fee_paid, discount_applied, total_paid, status,
                auto_renew, renewal_reminder_sent, total_freeze_days, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,0,0,?,?,0,0,0,?,?)
        """, (sub_id, user_id, plan_id, GYM_ID,
              start_date.isoformat(), end_date.isoformat(),
              total, total, status, ahora, ahora))
    return sub_id, status

def create_group_members(sub_id, primary_user_id, additional_user_ids):
    """Crea registros en subscription_members para el titular y adicionales."""
    ahora = now_iso()
    all_members = [(primary_user_id, True)] + [(uid, False) for uid in additional_user_ids]
    for user_id, is_primary in all_members:
        if not DRY_RUN:
            cur.execute("""
                INSERT INTO subscription_members (id, subscription_id, user_id, is_primary, created_at)
                VALUES (?,?,?,?,?)
            """, (new_uuid(), sub_id, user_id, 1 if is_primary else 0, ahora))

def has_active_sub(user_id):
    cur.execute("""SELECT id FROM subscriptions WHERE user_id=? AND status='ACTIVE' AND end_date > ?""",
                (user_id, datetime.now().isoformat()))
    if cur.fetchone(): return True
    # También revisar membresía grupal
    cur.execute("""
        SELECT sm.id FROM subscription_members sm
        JOIN subscriptions s ON s.id = sm.subscription_id
        WHERE sm.user_id=? AND s.status='ACTIVE' AND s.end_date > ?
    """, (user_id, datetime.now().isoformat()))
    return cur.fetchone() is not None

# ── PROCESAMIENTO ─────────────────────────────────────────────────────────────
print(f"{'[DRY RUN] ' if DRY_RUN else ''}DB: {DB_PATH}")
print()

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
stats = {"usuarios_nuevos": 0, "ya_existian": 0,
         "subs_creadas": 0, "omitidos": 0, "errores": 0, "sanitizados": 0}

# ═══════════════════════════════════════════════════════
# HOJA 1 — REGISTRO (individuales)
# ═══════════════════════════════════════════════════════
print("=" * 60)
print("REGISTRO (individuales)")
print("=" * 60)

for i, row in enumerate(wb["REGISTRO"].iter_rows(values_only=True)):
    if i < 4: continue
    try:
        nombre_raw  = row[0]
        cedula_raw  = row[1]
        telefono_raw= row[2]
        plan_raw    = row[3]
        fecha_raw   = row[4]
        dias_raw    = row[5]
        valor_raw   = row[8] if len(row) > 8 else None

        if not nombre_raw or not cedula_raw: continue

        nombre = clean_name(str(nombre_raw))
        doc    = clean_doc(cedula_raw)
        phone  = clean_phone(telefono_raw)
        parts  = nombre.split()
        first  = parts[0] if parts else "Sin"
        last   = " ".join(parts[1:]) if len(parts) > 1 else "Nombre"

        # Sanitizar doc con problemas
        if not doc or len(doc) < 4:
            print(f"  ! Cedula invalida [{cedula_raw}] para {nombre} — omitido")
            stats["errores"] += 1
            continue

        plan_key  = str(plan_raw).strip().upper() if plan_raw else "MENSUAL"
        plan_norm = PLAN_ALIAS.get(plan_key, "MENSUAL")
        plan_id   = get_or_create_plan(plan_norm)

        start = to_date(fecha_raw)
        days  = int(dias_raw) if dias_raw else 30
        price = float(valor_raw) if valor_raw else 0

        user_id, created = get_or_create_user(first, last, doc, phone)
        stats["usuarios_nuevos" if created else "ya_existian"] += 1

        if has_active_sub(user_id) and not created:
            stats["omitidos"] += 1
            continue

        sub_id, status = create_subscription(user_id, plan_id, start, days, price)
        stats["subs_creadas"] += 1
        icon = "🟢" if status == "ACTIVE" else "⚪"
        print(f"  {icon} {nombre} | {doc} | {plan_norm} {days}d | {start.strftime('%d/%m/%Y') if start else '?'} | ${price:,.0f}")

    except Exception as e:
        stats["errores"] += 1
        print(f"  ERROR fila {i+5}: {e}")

# ═══════════════════════════════════════════════════════
# HOJA 2 — PLANES ESPECIALES (grupos)
# ═══════════════════════════════════════════════════════
print()
print("=" * 60)
print("PLANES ESPECIALES (grupos)")
print("=" * 60)

for i, row in enumerate(wb["PLANES ESPECIALES"].iter_rows(values_only=True)):
    if i < 4: continue
    try:
        nombres_raw = row[0]
        cedulas_raw = row[1]
        tels_raw    = row[2]
        plan_raw    = row[3]
        fecha_raw   = row[4]
        dias_raw    = row[5]
        valor_raw   = row[8] if len(row) > 8 else None

        if not nombres_raw: continue

        nombres   = [clean_name(n) for n in split_multiline(nombres_raw) if n.strip()]
        cedulas   = [clean_doc(c)  for c in split_multiline(cedulas_raw)  if c.strip()]
        telefonos = [clean_phone(t) for t in split_multiline(tels_raw)    if t.strip()]

        if not nombres: continue

        plan_key  = str(plan_raw).strip().upper() if plan_raw else "PAREJA"
        plan_norm = PLAN_ALIAS.get(plan_key, plan_key)
        plan_id   = get_or_create_plan(plan_norm)

        start      = to_date(fecha_raw)
        days       = int(dias_raw) if dias_raw else 30
        price_total= float(valor_raw) if valor_raw else 0

        # Alinear nombres y cédulas (puede haber mismatch)
        if len(nombres) != len(cedulas):
            stats["sanitizados"] += 1
            # Rellenar cédulas faltantes con vacío
            while len(cedulas) < len(nombres): cedulas.append("")
            nombres = nombres[:len(cedulas)]

        icon_estado = "🟢" if start and add_duration(start, days) > datetime.now() else "⚪"
        print(f"\n  {icon_estado} {plan_norm} | {start.strftime('%d/%m/%Y') if start else '?'} | ${price_total:,.0f} | {len(nombres)} personas")

        user_ids = []
        for j, nombre in enumerate(nombres):
            doc   = cedulas[j] if j < len(cedulas) else ""
            phone = telefonos[j] if j < len(telefonos) else ""
            parts = nombre.split()
            first = parts[0] if parts else "Sin"
            last  = " ".join(parts[1:]) if len(parts) > 1 else "Nombre"

            if not doc or len(doc) < 4:
                # Sin cédula — registrar con doc placeholder
                doc = f"GRUP{new_uuid()[:6].upper()}"
                stats["sanitizados"] += 1
                print(f"    ! {nombre} sin cedula — asignado ID temporal {doc}")

            uid, created = get_or_create_user(first, last, doc, phone)
            stats["usuarios_nuevos" if created else "ya_existian"] += 1
            user_ids.append(uid)
            rol = "titular" if j == 0 else "adicional"
            print(f"    {'*' if j==0 else '-'} {nombre} | {cedulas[j] if j < len(cedulas) else '?'} [{rol}]")

        if not user_ids: continue

        primary_id      = user_ids[0]
        additional_ids  = user_ids[1:]

        if has_active_sub(primary_id):
            stats["omitidos"] += 1
            print(f"    ⏭  Titular ya tiene suscripcion activa — omitido")
            continue

        # Crear UNA sola suscripción para el titular con el precio total
        sub_id, status = create_subscription(primary_id, plan_id, start, days, price_total, is_primary=True)
        stats["subs_creadas"] += 1

        # Registrar todos los miembros del grupo (incluyendo titular)
        if additional_ids:
            create_group_members(sub_id, primary_id, additional_ids)

    except Exception as e:
        stats["errores"] += 1
        print(f"  ERROR fila {i+5}: {e}")

# ── Guardar ───────────────────────────────────────────────────────────────────
if not DRY_RUN:
    con.commit()
    print("\nCambios guardados en la base de datos.")
else:
    print("\n[DRY RUN] Sin cambios.")

con.close()

print()
print("=" * 60)
print("RESUMEN")
print("=" * 60)
print(f"  Usuarios nuevos:          {stats['usuarios_nuevos']}")
print(f"  Usuarios ya existian:     {stats['ya_existian']}")
print(f"  Suscripciones creadas:    {stats['subs_creadas']}")
print(f"  Registros sanitizados:    {stats['sanitizados']}")
print(f"  Omitidos (ya activos):    {stats['omitidos']}")
print(f"  Errores:                  {stats['errores']}")
